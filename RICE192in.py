# -*- coding: utf-8 -*-
"""
Created on Fri May 21 02:41:02 2021

@author: backm
"""

from subprocess import run
import glob
from datetime import datetime
import pandas as pd
import numpy as np
import openpyxl


def get_metadata(sheet_inp_sim):
    sim_label = compose_just_title(sheet_inp_sim)  # sheet_inp_sim["D2"].value
    sim_start = sheet_inp_sim["D4"].value.strftime("%d/%m/%Y")
    sim_end = sheet_inp_sim["D5"].value.strftime("%d/%m/%Y")
    meteo_code = sheet_inp_sim["D6"].value

    return meteo_code, sim_start, sim_end, sim_label


# =============================================================================
# def find_meteo(met_code):
#     """
#     Loading concrete meteo file donloaded from
#     https://eportal.mapa.gob.es/websiar/SeleccionParametrosMap.aspx?dst=1
#     """
#
#     met_code = met_code.lower()
#
#     meteos = glob.glob("meteo_data\\" + "*.csv")
#     codes = []
#     for meteo in meteos:
#         codes.append(meteo.split("\\")[-1].split("_")[0].lower())
#
#     df_meteo = pd.read_csv(meteos[codes.index(
#         met_code)], sep=";", decimal=",", header=0, encoding='UTF-16 LE')
#     return df_meteo
# =============================================================================

def find_meteo(met_code):
    """
    Loading concrete meteo file donloaded from
    https://eportal.mapa.gob.es/websiar/SeleccionParametrosMap.aspx?dst=1
    """
    df_meteo = pd.DataFrame()

    meteos = glob.glob("meteo_data\\" + "*.*")
    codes = []
    for meteo in meteos:
        codes.append(meteo.split("\\")[-1].split("_")[0].lower())
    met_code = met_code.lower()
    if met_code[0] == "v":
        df_meteo = pd.read_csv(meteos[codes.index(
            met_code)], sep=";", decimal=",", header=0, encoding='UTF-16 LE')
    elif met_code[0] == "a":
        df_meteo = pd.read_excel(meteos[codes.index(
            met_code)], index_col=0)
        df_meteo.Fecha = df_meteo.Fecha.apply(lambda x: x.strftime('%d/%m/%Y'))
    return df_meteo


def date_formater(date):
    date = date.split("/")
    dia = date[0]
    mes = date[1]
    anio = date[2][2:]

    if len(dia) == 1:
        dia = "0"+dia
    if mes[0] == "0":
        mes = mes[1:]
    date = mes+dia+anio
    return date.rjust(7)


def date_export(met_code, start_date="", end_date=""):
    """
    formato
    primer campo: fecha de la forma  7 espacios, mddaa
    segundo campo: precipitación 10 espacios, dividir por 10, 3 decimales
    tercer campo: evaporación 10 espacios, dividir por 10, 3 decimales
    cuarto campo: temperatura 10 espacios, 3 decimales
    """

    met = find_meteo(met_code)
    met["c1"] = met.Fecha.apply(date_formater)
    met['Fecha'] = pd.to_datetime(met['Fecha'], format='%d/%m/%Y')

    if start_date == "":
        start_date = met["Fecha"].min()

    if end_date == "":
        end_date = met["Fecha"].max()
    met = met[(met["Fecha"] >= pd.to_datetime(start_date, format='%d/%m/%Y'))
              & (met["Fecha"] <= pd.to_datetime(end_date, format='%d/%m/%Y'))]

    met["c2"] = met['Precipitación (mm)'].apply(
        lambda x: "{:9.3f}".format(x/10))
    met["c3"] = met['EtPMon'].apply(lambda x: "{:9.3f}".format(x/10))
    met["c4"] = met['Temp Media (ºC)'].apply(lambda x: "{:9.3f}".format(x))
    met_str = met.to_string(
        columns=["c1", "c2", "c3", "c4"], header=False, index=False, index_names=False)

    with open("bin\\ricewq.met", "w") as text_file:
        text_file.write(met_str)


def tst():
    result = run(r'RICE192.EXE', cwd=r'bin', shell=True,
                 universal_newlines=True,               capture_output=True)
    return result.stderr == 'RICEWQ TERMINATED PROPERLY\n'


def compose_just_title(sheet):
    cellname = sheet.cell(2, 4).value
    sheetname = sheet.title
    if sheetname[0] == "+":
        sheetname=sheetname[1:]
    if cellname == sheetname:
        sim_label = sheet.cell(2, 4).value
    else:
        sim_label = "{0}({1})".format(sheetname, cellname)
    return sim_label


def compose_title(sheet):
    sim_label = compose_just_title(sheet)

    sim_description = ["",""]

    cell_description = sheet.cell(3, 4).value
    cell_description=cell_description.replace("\n", " ")
    if len(cell_description) > 0:
        if len(cell_description) < 76:
            sim_description[0] = cell_description
        else:
            sim_description[0] = cell_description[:76]
            sim_description[1] = cell_description[76:]
    desc = """ {0}
     {1}""".format(sim_description[0], sim_description[1])


    return sim_label+"\n "+desc

def compose_parse_date_old(cell, leading_zero_month=False):

    try:
        dt = cell.value
    except:
        dt = cell
    day = str(dt.day)
    if len(day) == 1:
        day = "0"+day
    month = str(dt.month)
    if leading_zero_month:
        if len(month) == 1:
            month = "0"+month
    year = str(dt.year)[2:]
    return day, month, year

def compose_parse_date(cell, leading_zero_month=False):


    if isinstance(cell,openpyxl.cell.cell.Cell):
        dt = cell.value
    else:
        dt = cell

    if isinstance(dt,str):
        dt = datetime.strptime(dt,"%d/%m/%Y")

    day = str(dt.day)
    if len(day) == 1:
        day = "0"+day
    month = str(dt.month)
    if leading_zero_month:
        if len(month) == 1:
            month = "0"+month
    year = str(dt.year)[2:]
    return day, month, year


def compose_line_simdates(sheet):
    day_s, mon_s, year_s = compose_parse_date(sheet.cell(4, 4), True)
    day_f, mon_f, year_f = compose_parse_date(sheet.cell(5, 4), True)
    exfl = sheet.cell(7, 4).value
    template = "{0}  {1}  {2}   {3}  {4}   {5}    24    {6}".format(
        mon_s, day_s, year_s, mon_f, day_f, year_f, exfl)
    return template


def compose_line_cropdates(sheet):

    day_e, mon_e, year_foo = compose_parse_date(sheet.cell(8, 4), True)
    day_m, mon_m, year_foo = compose_parse_date(sheet.cell(9, 4), True)
    day_h, mon_h, year_foo = compose_parse_date(sheet.cell(10, 4), True)
    covmax = sheet.cell(11, 4).value
    ihfl = sheet.cell(12, 4).value
    template = "{0}  {1}    {2}  {3}   {4}   {5}    {6:0.2f}   {7}".format(
        mon_e, day_e, mon_m, day_m, mon_h, day_h, covmax, ihfl)
    return template


def compose_panel5B(sheet):
    cells = sheet["D"][12:22]
    template = "{0:4.2f}  {1:4.1f}   {2:4.1f}  {3:4.2f}  {4:4.1f}  {5:4.2f}  {6:4.2f} {7:4.2f}  {8:4.2f}   {9:4.1f}".format(
        *[cell.value for cell in cells])
    return template


def compose_panel12B(sheet):
    cells = sheet["D"][22:]
    template = "{0:d}    {1:d}     {2:d}     {3:d}    {4:d}      {5:d}      {6:d}  {7:5.2f}    {8:d}      {9:d}      {10:d}  {11:05.2f}".format(
        *[cell.value for cell in cells])
    return template


def max_decimal_in_string(str_in, num_max_dec=5):
    """ this function max out the number of decimals to a string """
    tmp_str = "{0:>.4f}".format(str_in)
    if tmp_str[0]!="0":
        tmp_str = tmp_str[:num_max_dec]
    else:
        tmp_str = tmp_str[1:num_max_dec+1]
    return tmp_str

def compose_hidrodates(sheet_inp_hidro):
    """ this function parses xlsx file for watering days """

    columns = ['date_irrigate', 'IRFLAG', 'DIRR1',
               'DIRR2', 'IRATE', 'DOUT', 'DR8MAX']
    pd_hidro = pd.DataFrame(sheet_inp_hidro.values)

    # Select only relevant columns
    pd_hidro = pd_hidro.iloc[: , :7]

    pd_hidro.drop([0, 1], axis=0, inplace=True)
    pd_hidro.columns = columns
    pd_hidro = pd_hidro.infer_objects()

    pd_hidro["date_irrigate"] = pd.to_datetime(pd_hidro.date_irrigate)

    str_out = str(len(pd_hidro))
    for row in pd_hidro.iterrows():
        row = row[1]
        day, month, year = compose_parse_date(
            row["date_irrigate"])
        line = """\n    {0:>2}  {1}     {2}   {3:4.1f}   {4:4.1f}  {5}  {6:4.1f}  {7}""".format(
            month, day, row["IRFLAG"], row["DIRR1"], row["DIRR2"], max_decimal_in_string(row["IRATE"]), row["DOUT"], max_decimal_in_string(row["DR8MAX"]))

        # Old line without maxin the decimal points
        # line = """\n    {0:>2}  {1}     {2}   {3:4.1f}   {4:4.1f}   {5:4.1f}  {6:4.1f}   {7:4.1f}""".format(
        #     month, day, row["IRFLAG"], row["DIRR1"], row["DIRR2"], row["IRATE"], row["DOUT"], row["DR8MAX"])
        str_out = str_out+line
    return str_out


def compose_applications(pd_applications):
    """ this function parses dataframe of applications"""

    str_out = str(len(pd_applications))
    for row in pd_applications.iterrows():
        row = row[1]

        try:
            day, month, year = compose_parse_date(
                row["date_pest_app"], leading_zero_month=True)
        except:
            day, month, year = compose_parse_date(
                pd.to_datetime(row["date_pest_app"]), leading_zero_month=True)
        line = """\n    {0}  {1}   {2:4.2f}  {3:4.2f}   {4:4.2f}   {5:3.1f}""".format(
            month, day, row["APP"], row["DINC"], row["APPEF"], row["DRIFT"])
        str_out = str_out+line

    return str_out


def compose_panel9B(pd_chemlist):
    """ this function parses dataframe of applications"""
    str_out = ""
    for row in pd_chemlist.iterrows():
        row = row[1]
        line = '\n{0: <9}   {1:3.1f}  {2:3.1f}  {3:3.1f}'.format(
            "\""+row["CNAME"]+"\"", row["CW0"], row["CS0"], row["CF0"])
        str_out = str_out+line

    return str_out[1:]


def rlz(string_in):
    """remove_leading_zero"""
    if string_in[0] == "0":
        string_out = string_in[1:]
    else:
        string_out = string_in
    return string_out


def compose_panel10B(pd_chemlist):
    """ this function parses dataframe of applications"""
    columns = ['KWM',
               'KWH',
               'KWP',
               'KSW',
               'KSD',
               'KF',
               'WO',
               'KD',
               'VVOL',
               'VSETL',
               'VBIND',
               'VMIX',
               'SOLUB',
               'RREAC',
               'SNK',
               'BI-P']
    str_out = ""
    for row in pd_chemlist.iterrows():
        row = row[1]
        processing_line = "{0:4.3f} {1:4.3f} {2:4.3f} {3:4.3f} {4:4.3f} {5:4.3f} {6:3.1f} {7:4.1f} {8:3.1f} {9:3.1f} {10:3.1f} {11:4.3f} {12:.1e} {13:4.2f} {14:3.1f} {15:1d}".format(
            *[row[col] for col in columns])
        processing_line = processing_line.split()
        processed_row = []
        for i, element in enumerate(processing_line):
            rlz_trigger = [0, 1, 2, 3, 4, 5, 11]
            if i in rlz_trigger:
                element = rlz(element)
            if i == 12:
                mantisa = element.split("+")[0].split("-")[0].upper()
                try:
                    exponente = rlz(element.split("+")[1])
                except:
                    exponente = rlz(element.split("-")[1])
                element = mantisa+exponente
            processed_row.append(element)
        new_line = "\n  {0} {1} {2}   {3}   {4} {5}  {6}  {7}  {8}   {9}   {10}  {11}  {12}  {13}  {14}   {15}".format(
            *[element for element in processed_row])
        str_out = str_out+new_line

    return str_out[1:]


def compose_panel11B(pd_chemlist):
    """ this function parses dataframe of applications"""
    columns = ['Q10WM', 'Q10SW', 'Q10SD', 'T1WM', 'T1SW', 'T1SD']
    str_out = ""
    for row in pd_chemlist.iterrows():
        row = row[1]
        line = "\n      {0:3.1f}   {1:3.1f}    {2:3.1f}   {3:4.1f}   {4:4.1f}   {5:4.1f}".format(
            *[row[col] for col in columns])
        str_out = str_out+line
    return str_out[1:]


def compose_chem_parents(pd_parents):
    """ this function parses dataframe of applications"""

    "      3         4 0.7  0.0  0.0  0.7 0.7  0.7"
    columns = ['PARENT',
               'DAUGHTER',
               'YWM(I)',
               'YWH(I)',
               'YWP(I)',
               'YSW(I)',
               'YSD(I)',
               'YF(I)']
    str_out = ""
    for row in pd_parents.iterrows():
        row = row[1]
        line = "\n      {0:1d}         {1:1d} {2:3.1f}  {3:3.1f}  {4:3.1f}  {5:3.1f} {6:3.1f}  {7:3.1f}".format(
            *[row[col] for col in columns])
        str_out = str_out+line
    return str_out[1:]


def parse_chem(sheet_inp_chem):
    """Function to parse a sheet of inp_chem.xlsx"""

    """Processing number of applications"""
    n_applications = sheet_inp_chem.cell(2, 4).value
    pd_chem = pd.DataFrame(sheet_inp_chem.values)
    pd_chem.drop([0], axis=1, inplace=True)
    pd_applications = pd_chem[2:7].set_index(1, drop=True).T.drop(
        index=[2]).reset_index(drop=True).infer_objects()
    pd_applications = pd_applications[0:n_applications]

    chem_applications = compose_applications(pd_applications)

    nchem = sheet_inp_chem["D8"].value
    npaths = sheet_inp_chem["D9"].value
    q10fl = sheet_inp_chem["D10"].value

    pd_chemlist = pd_chem[10:36].drop(
        2, axis=1).set_index(1, drop=True).T[:nchem]

    panel9b = compose_panel9B(pd_chemlist)
    panel10b = compose_panel10B(pd_chemlist)
    panel11b = compose_panel11B(pd_chemlist)

    pd_parents = pd_chem[36:44].drop(2, axis=1).set_index(
        1, drop=True).T.dropna().reset_index(drop=True)

    chem_parents = compose_chem_parents(pd_parents)
    template_inp_chem = """
** IAM IAD    APP  DINC  APPEF  DRIFT -- Appl. rates, efficiency (fraction), incorporation, drift(%)
    {0}
**  NCHEM  NPATHS   Q10FL
        {1:1d}       {2:1d}       {3:1d}
**  CNAME   CW0  CS0  CF0
{4}
** KWM  KWH  KWP    KSW    KSD   KF   WO   KD VVOL VSETL VBIND  VMIX SOLUB RREAC  SNK BI-P
{5}
**  Q10WM  Q10SW  Q10SD  T1WM  T1SW  T1SD
{6}
**PARENT DAUGHTER YWM  YWH  YWP  YSW YSD   YF  YWM2  YWH2  YWP2  YSW2  YSD2  YF2  JDAT
{7}
""".format(chem_applications, nchem, npaths, q10fl, panel9b, panel10b, panel11b, chem_parents)

    return template_inp_chem


def create_inp_file(sheet_inp_sim, sheet_inp_hidro, sheet_inp_chem):

    title = compose_title(sheet_inp_sim)
    simdates = compose_line_simdates(sheet_inp_sim)
    cropdates = compose_line_cropdates(sheet_inp_sim)

    template_inp_sim1 = """
 {0}
**  JM  JD  JY    KM  KD  KY   NTSD EXFL -- Simulation dates / EXAMS flag
    {1}
** JEM JED   KMM KMD  KHM  KHD  COVMAX IHFL -- Crop dates, canopy
    {2}
""".format(title, simdates, cropdates)

    hidrodates = compose_hidrodates(sheet_inp_hidro)

    template_inp_hidro = """
** IDM IDD IRFLAG  DIRR1  DIRR2 IRATE  DOUT DR8MAX  -- Irrigation & drainage
     {0}
""".format(hidrodates)

    panel5b = compose_panel5B(sheet_inp_sim)

    template_inp_sim2 = """
**  SA  DMAX  DLAKE  SEEP  DACT   FC   WP  SM    BD   CSS
  {0}
** EVAPM(I),I=-1
    -8.00  8.00  8.00 10.00 12.00 13.00 13.00 12.00 11.00 10.00  8.00  8.00
""".format(panel5b)

    template_inp_chem = parse_chem(sheet_inp_chem)

    panel12b = compose_panel12B(sheet_inp_sim)

    template_inp_sim3 = """
*** ENV CHM1 PRNT2 PRNT3 CHM2 NPROC2 RFORM2 YIELD2 CHN3 NPROC3 RFORM3 YIELD3
      {0}
*** END OF DATA
""".format(panel12b)

    template = template_inp_sim1[1:]+template_inp_hidro[1:] + \
        template_inp_sim2[1:]+template_inp_chem[1:] + template_inp_sim3[1:]

    with open("bin\\ricewq.inp", "w") as text_file:
        text_file.write(template)
