# -*- coding: utf-8 -*-
"""
Created on Wed May 19 19:38:52 2021

@author: backm
"""
import sys
from subprocess import run
import logging
from time import sleep
import openpyxl

import RICE192out as rout
from RICE192in import create_inp_file, date_export, get_metadata

### URL meteodata: https://eportal.mapa.gob.es/websiar/SeleccionParametrosMap.aspx?dst=1

# Console arguemnts
data = []
for i, arg in enumerate(sys.argv):
    if i == 0:
        pass
    else:
        data.append(sys.argv[i])



def main():
    inp_sim = openpyxl.load_workbook('input\inp_sim.xlsx', data_only=True)
    inp_sim_sheets = inp_sim.sheetnames

    inp_hidro = openpyxl.load_workbook('input\inp_hidro.xlsx', data_only=True)
    inp_hidro_sheets = inp_hidro.sheetnames

    inp_chem = openpyxl.load_workbook('input\inp_chem.xlsx', data_only=True)
    inp_chem_sheets = inp_chem.sheetnames

    # for sheet in inp_sim.sheetnames:
    sim_num = 1
    sim_tot = sum([1 if x[0] == "+" else 0 for x in inp_sim_sheets])*sum([1 if x[0] ==
                                                                          "+" else 0 for x in inp_hidro_sheets])*sum([1 if x[0] == "+" else 0 for x in inp_chem_sheets])
    for sim_sheet in inp_sim_sheets:
        if sim_sheet[0] != "+":
            continue
        sheet_inp_sim = inp_sim[sim_sheet]

        for hidro_label in inp_hidro_sheets:
            if hidro_label[0] != "+":
                continue

            # for sheet in inp_hidro.sheetnames:
            sheet_inp_hidro = inp_hidro[hidro_label]

            for chem_label in inp_chem_sheets:
                if chem_label[0] != "+":
                    continue
                # for sheet in inp_chem.sheetnames:
                sheet_inp_chem = inp_chem[chem_label]

                logging.basicConfig(level=logging.DEBUG, filename='bin/run.log')
                try:
                    meteo_code, sim_start, sim_end, sim_label = get_metadata(
                        sheet_inp_sim)
                    sim_name = "{0}_{1}_{2}".format(
                        sim_label, hidro_label[1:], chem_label[1:])
                    print("Starting {0}/{1}: {2}".format(sim_num, sim_tot, sim_name))

                    date_export(meteo_code, sim_start, sim_end)
                    create_inp_file(sheet_inp_sim, sheet_inp_hidro, sheet_inp_chem)

                    result = run(r'RICE192.EXE', cwd=r'bin', shell=True,
                                 universal_newlines=True, capture_output=True)
                    print(result.stderr)

                    if "error" in result.stderr.lower():
                        # Ex = ValueError()
                        # Ex.strerror = result.stderr
                        raise Exception('RICE192 EXCEPTION', result.stderr)

                    sleep(1)
                    rout.save_sim(sim_name=sim_name)
                    print("Saved.\n")
                except:
                    logging.exception("ERROR:")
                    sleep(1)
                    rout.move_results(sim_name="ERROR-"+sim_name)

                    if ERROR_BREAK:
                        sys.exit(1)
                    pass
                finally:
                    sim_num += 1


if __name__ == "__main__":

    ### STOP ON ERRORS
    ERROR_BREAK = True

    if len(data) > 0:
        if data[0].lower() == "n":
            ERROR_BREAK = False
    main()
